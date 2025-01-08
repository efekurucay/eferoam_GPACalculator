from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, make_response, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import math
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO
import os
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
import json
import numpy as np

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///gpa_calculator.db'
db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('Bu sayfaya erişim yetkiniz yok.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_number = db.Column(db.String(20), unique=True, nullable=False)
    first_name = db.Column(db.String(50), nullable=False)
    last_name = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128))
    is_admin = db.Column(db.Boolean, default=False)
    department = db.Column(db.String(10), nullable=False)  # Örn: CSE, PSI
    courses = db.relationship('Course', backref='user', lazy=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class CommonCourse(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    credit = db.Column(db.Float, nullable=False)
    akts = db.Column(db.Integer, nullable=False)
    semester_type = db.Column(db.String(20), nullable=False)  # "1-guz", "1-bahar", "2-guz", "2-bahar" vb.
    department = db.Column(db.String(10), nullable=False)  # Örn: CSE, PSI
    midterm_percentage = db.Column(db.Float, nullable=False)
    others_percentage = db.Column(db.Float, nullable=False)
    final_percentage = db.Column(db.Float, nullable=False)

class Course(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    common_course_id = db.Column(db.Integer, db.ForeignKey('common_course.id'), nullable=True)
    semester_type = db.Column(db.String(20), nullable=False)
    code = db.Column(db.String(20), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    credit = db.Column(db.Float, nullable=False)
    akts = db.Column(db.Integer, nullable=False)
    midterm_percentage = db.Column(db.Float, nullable=False)
    others_percentage = db.Column(db.Float, nullable=False)
    final_percentage = db.Column(db.Float, nullable=False)
    midterm_grade = db.Column(db.Float)
    others_grade = db.Column(db.Float)
    final_grade = db.Column(db.Float)
    total_grade = db.Column(db.Float)
    grade = db.Column(db.String(2))
    contribution_point = db.Column(db.Float)

    def update_grades(self, midterm=None, others=None, final=None):
        if midterm is not None:
            self.midterm_grade = float(midterm)
        if others is not None:
            self.others_grade = float(others)
        if final is not None:
            self.final_grade = float(final)

        if any(grade is not None for grade in [self.midterm_grade, self.others_grade, self.final_grade]):
            total = 0
            if self.midterm_grade is not None:
                total += self.midterm_grade * (self.midterm_percentage / 100)
            if self.others_grade is not None:
                total += self.others_grade * (self.others_percentage / 100)
            if self.final_grade is not None:
                total += self.final_grade * (self.final_percentage / 100)
            
            self.total_grade = math.floor(total + 0.5)

            if self.total_grade >= 88:
                self.grade = 'AA'
                self.contribution_point = 4.0
            elif self.total_grade >= 81:
                self.grade = 'BA'
                self.contribution_point = 3.5
            elif self.total_grade >= 74:
                self.grade = 'BB'
                self.contribution_point = 3.0
            elif self.total_grade >= 67:
                self.grade = 'CB'
                self.contribution_point = 2.5
            elif self.total_grade >= 60:
                self.grade = 'CC'
                self.contribution_point = 2.0
            elif self.total_grade >= 53:
                self.grade = 'DC'
                self.contribution_point = 1.5
            elif self.total_grade >= 46:
                self.grade = 'DD'
                self.contribution_point = 1.0
            else:
                self.grade = 'FF'
                self.contribution_point = 0.0

class Note(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    course_id = db.Column(db.Integer, db.ForeignKey('course.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    title = db.Column(db.String(100), nullable=False)
    content = db.Column(db.Text, nullable=False)
    due_date = db.Column(db.DateTime)
    is_reminder = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    course = db.relationship('Course', backref='notes')

class SharedNote(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    common_course_id = db.Column(db.Integer, db.ForeignKey('common_course.id'), nullable=False)
    admin_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    title = db.Column(db.String(100), nullable=False)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    common_course = db.relationship('CommonCourse', backref='shared_notes')
    admin = db.relationship('User', backref='created_shared_notes')

class SharedNoteAccess(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    shared_note_id = db.Column(db.Integer, db.ForeignKey('shared_note.id'), nullable=False)
    student_number = db.Column(db.String(20), nullable=False)
    granted_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    shared_note = db.relationship('SharedNote', backref='access_list')

def get_semester_display(semester_type):
    parts = semester_type.split('-')
    year = parts[0]
    term = parts[1]
    return f"{year}. Sınıf {term.capitalize()}"

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def calculate_gpa(courses):
    if not courses:
        return 0.0
    
    total_credit = 0
    total_point = 0
    
    for course in courses:
        if course.grade and course.contribution_point is not None:
            total_credit += course.credit
            total_point += course.credit * course.contribution_point
    
    return total_point / total_credit if total_credit > 0 else 0.0

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('index.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        student_number = request.form.get('student_number')
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        email = request.form.get('email')
        password = request.form.get('password')
        department = request.form.get('department')

        if User.query.filter_by(email=email).first():
            flash('Bu e-posta adresi zaten kayıtlı.', 'danger')
            return redirect(url_for('register'))

        if User.query.filter_by(student_number=student_number).first():
            flash('Bu öğrenci numarası zaten kayıtlı.', 'danger')
            return redirect(url_for('register'))

        user = User(
            student_number=student_number,
            first_name=first_name,
            last_name=last_name,
            email=email,
            department=department
        )
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash('Kayıt başarıyla tamamlandı. Giriş yapabilirsiniz.', 'success')
        return redirect(url_for('login'))

    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        student_number = request.form.get('student_number')
        password = request.form.get('password')
        
        user = User.query.filter_by(student_number=student_number).first()
        
        if user and user.check_password(password):
            login_user(user)
            next_page = request.args.get('next')
            flash('Başarıyla giriş yaptınız!', 'success')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        else:
            flash('Öğrenci numarası veya şifre hatalı!', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Başarıyla çıkış yaptınız.', 'success')
    return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    courses = Course.query.filter_by(user_id=current_user.id).all()
    
    # Benzersiz dönemleri al
    semesters = sorted(list(set(course.semester_type for course in courses)))
    
    # URL'den seçili dönemi al
    selected_semester = request.args.get('semester', 'all')
    
    return render_template('dashboard.html', 
                         courses=courses, 
                         calculate_gpa=calculate_gpa, 
                         get_semester_display=get_semester_display,
                         semesters=semesters,
                         selected_semester=selected_semester)

@app.route('/add_course', methods=['GET', 'POST'])
@login_required
def add_course():
    common_courses = CommonCourse.query.all()
    
    if request.method == 'POST':
        common_course_id = request.form.get('common_course_id')
        
        if common_course_id:
            common_course = CommonCourse.query.get(int(common_course_id))
            course = Course(
                user_id=current_user.id,
                common_course_id=common_course.id,
                semester_type=common_course.semester_type,
                code=common_course.code,
                name=common_course.name,
                credit=common_course.credit,
                akts=common_course.akts,
                midterm_percentage=common_course.midterm_percentage,
                others_percentage=common_course.others_percentage,
                final_percentage=common_course.final_percentage
            )
        else:
            course = Course(
                user_id=current_user.id,
                semester_type=request.form.get('semester_type'),
                code=request.form.get('code'),
                name=request.form.get('name'),
                credit=float(request.form.get('credit')),
                akts=int(request.form.get('akts')),
                midterm_percentage=float(request.form.get('midterm_percentage')),
                others_percentage=float(request.form.get('others_percentage')),
                final_percentage=float(request.form.get('final_percentage'))
            )
        
        db.session.add(course)
        db.session.commit()
        flash('Ders başarıyla eklendi.', 'success')
        return redirect(url_for('dashboard'))

    return render_template('add_course.html', common_courses=common_courses, get_semester_display=get_semester_display)

@app.route('/update_grades/<int:course_id>', methods=['GET', 'POST'])
@login_required
def update_grades(course_id):
    course = Course.query.get_or_404(course_id)
    if course.user_id != current_user.id:
        flash('Bu işlem için yetkiniz yok.', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        midterm = request.form.get('midterm_grade')
        others = request.form.get('others_grade')
        final = request.form.get('final_grade')
        
        course.update_grades(
            midterm=float(midterm) if midterm else None,
            others=float(others) if others else None,
            final=float(final) if final else None
        )
        
        db.session.commit()
        flash('Notlar başarıyla güncellendi.', 'success')
        return redirect(url_for('dashboard'))

    return render_template('update_grades.html', course=course, get_semester_display=get_semester_display)

@app.route('/admin/courses')
@login_required
@admin_required
def admin_courses():
    courses = CommonCourse.query.all()
    return render_template('admin/courses.html', courses=courses, get_semester_display=get_semester_display)

@app.route('/admin/add_course', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_add_course():
    if request.method == 'POST':
        course = CommonCourse(
            code=request.form.get('code'),
            name=request.form.get('name'),
            credit=float(request.form.get('credit')),
            akts=int(request.form.get('akts')),
            semester_type=request.form.get('semester_type'),
            department=request.form.get('department'),
            midterm_percentage=float(request.form.get('midterm_percentage')),
            others_percentage=float(request.form.get('others_percentage')),
            final_percentage=float(request.form.get('final_percentage'))
        )
        db.session.add(course)
        db.session.commit()
        flash('Ders başarıyla eklendi.', 'success')
        return redirect(url_for('admin_courses'))

    return render_template('admin/add_course.html')

@app.route('/admin/edit_course/<int:course_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_edit_course(course_id):
    course = CommonCourse.query.get_or_404(course_id)
    
    if request.method == 'POST':
        course.code = request.form.get('code')
        course.name = request.form.get('name')
        course.credit = float(request.form.get('credit'))
        course.akts = int(request.form.get('akts'))
        course.semester_type = request.form.get('semester_type')
        course.midterm_percentage = float(request.form.get('midterm_percentage'))
        course.others_percentage = float(request.form.get('others_percentage'))
        course.final_percentage = float(request.form.get('final_percentage'))
        
        db.session.commit()
        flash('Ders başarıyla güncellendi.', 'success')
        return redirect(url_for('admin_courses'))

    return render_template('admin/edit_course.html', course=course)

@app.route('/admin/delete_course/<int:course_id>', methods=['POST'])
@login_required
@admin_required
def admin_delete_course(course_id):
    course = CommonCourse.query.get_or_404(course_id)
    db.session.delete(course)
    db.session.commit()
    flash('Ders başarıyla silindi.', 'success')
    return redirect(url_for('admin_courses'))

@app.route('/delete_course/<int:course_id>', methods=['POST'])
@login_required
def delete_course(course_id):
    course = Course.query.get_or_404(course_id)
    if course.user_id != current_user.id:
        flash('Bu işlem için yetkiniz yok.', 'danger')
        return redirect(url_for('dashboard'))
    
    db.session.delete(course)
    db.session.commit()
    flash('Ders başarıyla silindi.', 'success')
    return redirect(url_for('dashboard'))

@app.route('/export/<format>')
@login_required
def export_grades(format):
    if format not in ['excel', 'pdf']:
        flash('Geçersiz dosya formatı.', 'danger')
        return redirect(url_for('dashboard'))
    
    courses = Course.query.filter_by(user_id=current_user.id).all()
    
    if not courses:
        flash('Dışa aktarılacak ders bulunamadı.', 'warning')
        return redirect(url_for('dashboard'))
    
    if format == 'excel':
        return export_excel(courses)
    else:
        return export_pdf(courses)

def export_excel(courses):
    wb = Workbook()
    ws = wb.active
    ws.title = "Not Dökümü"
    
    # Başlık stilleri
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    centered = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlıklar
    headers = [
        "Dönem", "Ders Kodu", "Ders Adı", "Kredi", "AKTS",
        "MT%", "Others%", "Final%", "Midterm", "Others", "Final",
        "Total", "Harf", "Katkı"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
        cell.border = border
    
    # Veriler
    for row, course in enumerate(sorted(courses, key=lambda x: x.semester_type), 2):
        data = [
            get_semester_display(course.semester_type),
            course.code,
            course.name,
            course.credit,
            course.akts,
            course.midterm_percentage,
            course.others_percentage,
            course.final_percentage,
            course.midterm_grade,
            course.others_grade,
            course.final_grade,
            course.total_grade,
            course.grade,
            course.contribution_point
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.alignment = centered
            cell.border = border
    
    # Sütun genişliklerini ayarla
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    # Özet bilgiler
    summary_row = len(courses) + 4
    ws.cell(row=summary_row, column=1, value="Genel Ortalama:")
    ws.cell(row=summary_row, column=2, value=calculate_gpa(courses))
    
    # Excel dosyasını kaydet
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"not_dokumu_{current_user.student_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

def export_pdf(courses):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Stil tanımlamaları
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30
    )
    
    # Başlık
    title = Paragraph(f"Not Dökümü - {current_user.first_name} {current_user.last_name}", title_style)
    elements.append(title)
    elements.append(Paragraph(f"Öğrenci Numarası: {current_user.student_number}", styles["Normal"]))
    elements.append(Spacer(1, 20))
    
    # Tablo verileri
    headers = [
        "Dönem", "Ders Kodu", "Ders Adı", "Kredi", "AKTS",
        "MT%", "Others%", "Final%", "Midterm", "Others", "Final",
        "Total", "Harf", "Katkı"
    ]
    
    data = [headers]
    
    for course in sorted(courses, key=lambda x: x.semester_type):
        row = [
            get_semester_display(course.semester_type),
            course.code,
            course.name,
            str(course.credit),
            str(course.akts),
            f"{course.midterm_percentage}%",
            f"{course.others_percentage}%",
            f"{course.final_percentage}%",
            str(course.midterm_grade) if course.midterm_grade else "-",
            str(course.others_grade) if course.others_grade else "-",
            str(course.final_grade) if course.final_grade else "-",
            str(course.total_grade) if course.total_grade else "-",
            course.grade if course.grade else "-",
            str(course.contribution_point) if course.contribution_point else "-"
        ]
        data.append(row)
    
    # Tablo stilleri
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Genel ortalama
    elements.append(Paragraph(f"Genel Ortalama: {calculate_gpa(courses):.2f}", styles["Normal"]))
    
    # PDF oluştur
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"not_dokumu_{current_user.student_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@app.route('/course/<int:course_id>/notes')
@login_required
def course_notes(course_id):
    course = Course.query.get_or_404(course_id)
    if course.user_id != current_user.id:
        flash('Bu işlem için yetkiniz yok.', 'danger')
        return redirect(url_for('dashboard'))
    
    notes = Note.query.filter_by(course_id=course_id).order_by(Note.created_at.desc()).all()
    return render_template('notes.html', course=course, notes=notes)

@app.route('/course/<int:course_id>/notes/add', methods=['GET', 'POST'])
@login_required
def add_note(course_id):
    course = Course.query.get_or_404(course_id)
    if course.user_id != current_user.id:
        flash('Bu işlem için yetkiniz yok.', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        note = Note(
            course_id=course_id,
            user_id=current_user.id,
            title=request.form.get('title'),
            content=request.form.get('content'),
            is_reminder=bool(request.form.get('is_reminder')),
            due_date=datetime.strptime(request.form.get('due_date'), '%Y-%m-%dT%H:%M') if request.form.get('due_date') else None
        )
        db.session.add(note)
        db.session.commit()
        flash('Not başarıyla eklendi.', 'success')
        return redirect(url_for('course_notes', course_id=course_id))
    
    return render_template('add_note.html', course=course)

@app.route('/note/<int:note_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_note(note_id):
    note = Note.query.get_or_404(note_id)
    if note.user_id != current_user.id:
        flash('Bu işlem için yetkiniz yok.', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        note.title = request.form.get('title')
        note.content = request.form.get('content')
        note.is_reminder = bool(request.form.get('is_reminder'))
        note.due_date = datetime.strptime(request.form.get('due_date'), '%Y-%m-%dT%H:%M') if request.form.get('due_date') else None
        db.session.commit()
        flash('Not başarıyla güncellendi.', 'success')
        return redirect(url_for('course_notes', course_id=note.course_id))
    
    return render_template('edit_note.html', note=note)

@app.route('/note/<int:note_id>/delete', methods=['POST'])
@login_required
def delete_note(note_id):
    note = Note.query.get_or_404(note_id)
    if note.user_id != current_user.id:
        flash('Bu işlem için yetkiniz yok.', 'danger')
        return redirect(url_for('dashboard'))
    
    course_id = note.course_id
    db.session.delete(note)
    db.session.commit()
    flash('Not başarıyla silindi.', 'success')
    return redirect(url_for('course_notes', course_id=course_id))

@app.route('/reminders')
@login_required
def reminders():
    upcoming_reminders = Note.query.filter_by(
        user_id=current_user.id,
        is_reminder=True
    ).filter(
        Note.due_date >= datetime.utcnow()
    ).order_by(
        Note.due_date
    ).all()
    
    return render_template('reminders.html', reminders=upcoming_reminders)

@app.route('/admin/courses/<int:course_id>/shared_notes')
@login_required
@admin_required
def admin_shared_notes(course_id):
    course = CommonCourse.query.get_or_404(course_id)
    shared_notes = SharedNote.query.filter_by(common_course_id=course_id).all()
    return render_template('admin/shared_notes.html', course=course, shared_notes=shared_notes)

@app.route('/admin/courses/<int:course_id>/shared_notes/add', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_add_shared_note(course_id):
    course = CommonCourse.query.get_or_404(course_id)
    
    if request.method == 'POST':
        note = SharedNote(
            common_course_id=course_id,
            admin_id=current_user.id,
            title=request.form.get('title'),
            content=request.form.get('content')
        )
        db.session.add(note)
        db.session.commit()
        
        # Erişim izinlerini ekle
        student_numbers = request.form.get('student_numbers', '').strip().split('\n')
        for student_number in student_numbers:
            student_number = student_number.strip()
            if student_number:  # Boş satırları atla
                access = SharedNoteAccess(
                    shared_note_id=note.id,
                    student_number=student_number
                )
                db.session.add(access)
        
        db.session.commit()
        flash('Ortak not başarıyla eklendi.', 'success')
        return redirect(url_for('admin_shared_notes', course_id=course_id))
    
    return render_template('admin/add_shared_note.html', course=course)

@app.route('/admin/shared_notes/<int:note_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_edit_shared_note(note_id):
    note = SharedNote.query.get_or_404(note_id)
    
    if request.method == 'POST':
        note.title = request.form.get('title')
        note.content = request.form.get('content')
        
        # Mevcut erişim izinlerini sil
        SharedNoteAccess.query.filter_by(shared_note_id=note.id).delete()
        
        # Yeni erişim izinlerini ekle
        student_numbers = request.form.get('student_numbers', '').strip().split('\n')
        for student_number in student_numbers:
            student_number = student_number.strip()
            if student_number:  # Boş satırları atla
                access = SharedNoteAccess(
                    shared_note_id=note.id,
                    student_number=student_number
                )
                db.session.add(access)
        
        db.session.commit()
        flash('Ortak not başarıyla güncellendi.', 'success')
        return redirect(url_for('admin_shared_notes', course_id=note.common_course_id))
    
    # Mevcut erişim izinlerini al
    current_access = SharedNoteAccess.query.filter_by(shared_note_id=note.id).all()
    student_numbers = '\n'.join(access.student_number for access in current_access)
    
    return render_template('admin/edit_shared_note.html', note=note, student_numbers=student_numbers)

@app.route('/admin/shared_notes/<int:note_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_shared_note(note_id):
    note = SharedNote.query.get_or_404(note_id)
    course_id = note.common_course_id
    
    # Erişim izinlerini sil
    SharedNoteAccess.query.filter_by(shared_note_id=note.id).delete()
    
    # Notu sil
    db.session.delete(note)
    db.session.commit()
    
    flash('Ortak not başarıyla silindi.', 'success')
    return redirect(url_for('admin_shared_notes', course_id=course_id))

@app.route('/course/<int:course_id>/shared_notes')
@login_required
def shared_notes(course_id):
    course = Course.query.get_or_404(course_id)
    if course.user_id != current_user.id:
        flash('Bu işlem için yetkiniz yok.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Ortak dersin paylaşılan notlarını bul
    if course.common_course_id:
        shared_notes = SharedNote.query.join(SharedNoteAccess).filter(
            SharedNote.common_course_id == course.common_course_id,
            SharedNoteAccess.student_number == current_user.student_number
        ).all()
    else:
        shared_notes = []
    
    return render_template('shared_notes.html', course=course, shared_notes=shared_notes)

@app.route('/statistics')
@login_required
def statistics():
    courses = Course.query.filter_by(user_id=current_user.id).all()
    
    if not courses:
        flash('İstatistikleri görüntülemek için en az bir ders eklemelisiniz.', 'warning')
        return redirect(url_for('dashboard'))
    
    # Bölüm ortalaması ve sıralama hesaplama
    department_users = User.query.filter_by(department=current_user.department).all()
    department_gpas = []
    for user in department_users:
        user_courses = Course.query.filter_by(user_id=user.id).all()
        if user_courses:
            gpa = calculate_gpa(user_courses)
            department_gpas.append({
                'gpa': gpa,
                'is_current_user': user.id == current_user.id
            })
    
    # GPA'lere göre sırala (büyükten küçüğe)
    department_gpas.sort(key=lambda x: x['gpa'], reverse=True)
    
    # Kullanıcının sıralamasını bul
    total_students = len(department_gpas)
    user_rank = next((i + 1 for i, item in enumerate(department_gpas) if item['is_current_user']), 0)
    percentile = round((total_students - user_rank + 1) / total_students * 100) if total_students > 0 else 0
    
    department_avg = sum(item['gpa'] for item in department_gpas) / len(department_gpas) if department_gpas else 0
    
    # Ders bazlı ortalama hesaplama
    course_averages = []
    for course in courses:
        if course.common_course_id:
            # Ortak dersin tüm öğrencilerdeki notlarını bul
            all_course_instances = Course.query.join(User).filter(
                Course.common_course_id == course.common_course_id,
                User.department == current_user.department
            ).all()
            
            # Notları sırala
            valid_instances = [
                {
                    'student_id': instance.user_id,
                    'point': instance.contribution_point
                }
                for instance in all_course_instances
                if instance.contribution_point is not None
            ]
            valid_instances.sort(key=lambda x: x['point'], reverse=True)
            
            # Öğrencinin sıralamasını bul
            total_students = len(valid_instances)
            student_rank = next((i + 1 for i, item in enumerate(valid_instances) if item['student_id'] == current_user.id), 0)
            student_percentile = round((total_students - student_rank + 1) / total_students * 100) if total_students > 0 else 0
            
            # Ortalama hesapla
            total_points = sum(instance['point'] for instance in valid_instances)
            course_avg = total_points / total_students if total_students > 0 else 0
            
            course_averages.append({
                'code': course.code,
                'name': course.name,
                'my_grade': course.grade,
                'my_point': course.contribution_point,
                'class_average': round(course_avg, 2),
                'total_students': total_students,
                'my_rank': student_rank,
                'percentile': student_percentile
            })
    
    # Dönem bazlı GPA grafiği
    semester_data = {}
    for course in courses:
        if course.semester_type not in semester_data:
            semester_data[course.semester_type] = []
        semester_data[course.semester_type].append(course)
    
    semester_gpas = []
    for semester, semester_courses in sorted(semester_data.items()):
        gpa = calculate_gpa(semester_courses)
        semester_gpas.append({
            'semester': get_semester_display(semester),
            'gpa': round(gpa, 2)
        })
    
    gpa_fig = {
        'data': [{
            'x': [d['semester'] for d in semester_gpas],
            'y': [d['gpa'] for d in semester_gpas],
            'type': 'scatter',
            'mode': 'lines+markers',
            'name': 'GPA'
        }],
        'layout': {
            'title': {
                'text': 'Dönem Bazlı GPA Değişimi',
                'x': 0.5,
                'font': {'size': 20}
            },
            'showlegend': False,
            'plot_bgcolor': 'white',
            'height': 400,
            'margin': {'l': 50, 'r': 50, 't': 50, 'b': 50},
            'xaxis': {'title': 'Dönem'},
            'yaxis': {'title': 'GPA'}
        }
    }
    
    # Harf notu dağılımı
    grade_counts = {}
    for course in courses:
        if course.grade:
            grade_counts[course.grade] = grade_counts.get(course.grade, 0) + 1
    
    grade_order = ['AA', 'BA', 'BB', 'CB', 'CC', 'DC', 'DD', 'FF']
    grade_colors = {
        'AA': '#28a745', 'BA': '#28a745', 'BB': '#28a745',
        'CB': '#17a2b8', 'CC': '#17a2b8',
        'DC': '#ffc107', 'DD': '#ffc107',
        'FF': '#dc3545'
    }
    
    grades_data = []
    colors = []
    for grade in grade_order:
        if grade in grade_counts:
            grades_data.append(grade_counts[grade])
            colors.append(grade_colors[grade])
    
    grades_fig = {
        'data': [{
            'x': [grade for grade in grade_order if grade in grade_counts],
            'y': grades_data,
            'type': 'bar',
            'marker': {'color': colors}
        }],
        'layout': {
            'title': {
                'text': 'Harf Notu Dağılımı',
                'x': 0.5,
                'font': {'size': 20}
            },
            'showlegend': False,
            'plot_bgcolor': 'white',
            'height': 400,
            'margin': {'l': 50, 'r': 50, 't': 50, 'b': 50},
            'xaxis': {'title': 'Harf Notu'},
            'yaxis': {'title': 'Ders Sayısı'}
        }
    }
    
    return render_template(
        'statistics.html',
        gpa_chart=json.dumps(gpa_fig),
        grades_chart=json.dumps(grades_fig),
        current_gpa=round(calculate_gpa(courses), 2),
        department_avg=round(department_avg, 2),
        total_credits=sum(course.credit for course in courses if course.grade),
        total_courses=len([course for course in courses if course.grade]),
        course_averages=course_averages,
        user_rank=user_rank,
        total_students=total_students,
        percentile=percentile
    )

@app.route('/sw.js')
def service_worker():
    response = make_response(send_from_directory('static', 'sw.js'))
    response.headers['Content-Type'] = 'application/javascript'
    return response

@app.route('/offline.html')
def offline():
    return render_template('offline.html')

@app.route('/manifest.json')
def manifest():
    return send_from_directory('static', 'manifest.json')

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True) 